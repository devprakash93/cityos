import io
from datetime import timedelta
from django.utils import timezone
from django.db.models import Count, Avg, Sum
from apps.complaints.models import Complaint

def generate_pdf_report(start_date, end_date, user):
    """Generate a PDF system performance report using ReportLab."""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"Odisha CityOS - System Performance Report", styles['Title']))
    story.append(Paragraph(f"Period: {start_date.date()} to {end_date.date()}", styles['Normal']))
    story.append(Spacer(1, 12))

    # Complaint Summary
    qs = Complaint.objects.filter(created_at__range=(start_date, end_date))
    
    if user.role and user.role.name == "OFFICER" and user.department:
        qs = qs.filter(department=user.department)
        
    from apps.geography.permissions import get_city_for_user
    city = get_city_for_user(user)
    if city:
        qs = qs.filter(ward__city=city)
        
    total_complaints = qs.count()
    resolved = qs.filter(status=Complaint.RESOLVED).count()
    breaches = qs.filter(sla_breached=True).count()
    resolution_rate = round(resolved / total_complaints * 100, 1) if total_complaints else 0

    story.append(Paragraph("1. Complaint SLA & Resolution Summary", styles['Heading2']))
    
    data = [
        ["Total Complaints", "Resolved", "Resolution Rate", "SLA Breaches"],
        [str(total_complaints), str(resolved), f"{resolution_rate}%", str(breaches)]
    ]
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(t)
    story.append(Spacer(1, 12))

    # By Department
    story.append(Paragraph("2. Department Breakdown", styles['Heading2']))
    dept_stats = qs.values("department__name").annotate(count=Count("id")).order_by("-count")
    
    dept_data = [["Department", "Complaints"]]
    for stat in dept_stats:
        dept_data.append([stat['department__name'], str(stat['count'])])
        
    t2 = Table(dept_data)
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(t2)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(start_date, end_date, user):
    """Generate an Excel export using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints Export"

    # Header
    headers = ["ID", "Citizen", "Department", "Category", "Status", "Priority", "Created At", "SLA Breached"]
    ws.append(headers)
    
    # Style Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    qs = Complaint.objects.filter(created_at__range=(start_date, end_date)).select_related("citizen", "department", "category")
    
    if user.role and user.role.name == "OFFICER" and user.department:
        qs = qs.filter(department=user.department)
        
    from apps.geography.permissions import get_city_for_user
    city = get_city_for_user(user)
    if city:
        qs = qs.filter(ward__city=city)
    
    for row_num, complaint in enumerate(qs, start=2):
        ws.cell(row=row_num, column=1, value=complaint.id)
        ws.cell(row=row_num, column=2, value=complaint.citizen.get_full_name() if complaint.citizen else "Unknown")
        ws.cell(row=row_num, column=3, value=complaint.department.name if complaint.department else "")
        ws.cell(row=row_num, column=4, value=complaint.category.name if complaint.category else "")
        ws.cell(row=row_num, column=5, value=complaint.status)
        ws.cell(row=row_num, column=6, value=complaint.priority)
        ws.cell(row=row_num, column=7, value=complaint.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=8, value="Yes" if complaint.sla_breached else "No")

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
